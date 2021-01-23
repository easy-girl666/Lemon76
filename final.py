import openpyxl
import requests
#读取数据
def readdata(filename,sheetname):
    wb=openpyxl.load_workbook(filename)
    sh=wb[sheetname]
    max_row=sh.max_row
    cases=[]
    for row in range(2,max_row+1):
        dict1 = dict(
        id = sh.cell(row=row, column=1).value,
        url = sh.cell(row=row,column=5).value,
        data = eval(sh.cell(row=row, column=6).value),                              #从Excel读取出来的数据是str
        expected = eval(sh.cell(row=row, column=7).value))
        cases.append(dict1)
    return cases

def func(url,data):
    header = {'X-Lemonban-Media-Type': 'lemonban.v2',
                  'Content-Type': 'application/json'}
    res1=requests.post(url=url, json=data, headers=header)
    res2=res1.json()
    return res2

def write(filename,sheetname,row,final_result):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    sh.cell(row=row,column=8).value=final_result
    wb.save(filename)
    wb.close()



def execute(filename,sheetname):
    res = readdata(filename, sheetname)
    # print(res)
    for case in res:
        expected=case['expected']
        caseid1=case['id']
        res3=func(case['url'],case['data'])
        # print(res3)
        print('第{}条用例预期结果是{}： '.format(caseid1,expected['msg']))
        print('第{}条用例实际结果是{}： '.format(caseid1, res3['msg']))
        if expected['msg'] == res3['msg']:
            final_result = 'pass'
            print('这条用例通过')
        else:
            final_result = 'fail'
            print('这条用例不通过')
        res4=write(filename, sheetname,caseid1+1,final_result)
        print(res4)

res5=execute('test_case_api.xlsx','login')
print(res5)

